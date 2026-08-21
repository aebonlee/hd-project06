# -*- coding: utf-8 -*-
"""
make_samples.py — 표준 엑셀 양식 / 샘플 데이터 생성 스크립트

산출물:
  1) templates/엔진사양_표준양식.xlsx : 빈 양식 (헤더 + schema 시트만 채움)
  2) templates/엔진사양_샘플.xlsx     : 제조사 3곳, 엔진 6개 샘플 데이터
  3) js/sample-data.js               : 동일 데이터를 웹페이지에 내장 ("샘플 데이터 불러오기" 버튼용)

곡선 데이터는 물리적으로 그럴듯하게 생성한다:
  - 토크는 중간 RPM 대역에서 최대(포물선 형태), 출력 = 토크 x RPM / 9549
  - 정격 출력 = 정격 RPM에서의 곡선상 출력값, 최대 토크 = 곡선상 최대 토크값
    (engines 시트의 정격/최대 값이 curves 시트와 반드시 일치 — 3.2.1.3 주석 라벨의 기준)

주의: js/data.js 의 DEFAULT_SCHEMA / 시트 이름 / 헤더와 완전히 동일해야 한다.
"""
import json
import os

from openpyxl import Workbook

BASE = os.path.dirname(os.path.abspath(__file__))

# ---------------------------------------------------------------- schema
# (항목 키, 표시명, 그룹, 단위, 순서) — js/data.js 의 DEFAULT_SCHEMA 와 동일
SCHEMA = [
    ("manufacturer", "엔진 제조사", "기본 정보", "", 10),
    ("engine_model", "엔진 모델명", "기본 정보", "", 20),
    ("forklift_models", "적용 지게차 모델", "기본 정보", "", 30),
    ("rated_power_kw", "정격 출력", "성능", "kW", 40),
    ("rated_power_rpm", "정격 출력 RPM", "성능", "rpm", 50),
    ("max_torque_nm", "최대 토크", "성능", "Nm", 60),
    ("max_torque_rpm", "최대 토크 RPM", "성능", "rpm", 70),
    ("low_idle_rpm", "Low Idle RPM", "성능", "rpm", 80),
    ("high_idle_rpm", "High Idle RPM", "성능", "rpm", 90),
    ("weight_kg", "무게", "주요 제원", "kg", 100),
    ("engine_type", "엔진 타입", "주요 제원", "", 110),
    ("cylinders", "실린더 수", "주요 제원", "개", 120),
    ("displacement_cc", "배기량", "주요 제원", "cc", 130),
    ("oil_capacity_l", "오일 용량", "주요 제원", "L", 140),
    ("coolant_capacity_l", "냉각수 용량", "주요 제원", "L", 150),
    ("fuel_consumption", "연비(정격점)", "주요 제원", "g/kWh", 160),
    ("alternator", "알터네이터 스펙", "주요 제원", "", 170),
    ("starter_motor_kw", "스타터 모터 출력", "주요 제원", "kW", 180),
    ("emission_cert", "배기 규제 인증", "규제·인증", "", 190),
    ("aftertreatment", "후처리장치 타입", "규제·인증", "", 200),
]
SCHEMA_HEADERS = ["항목 키", "표시명", "그룹", "단위", "순서"]
CURVE_HEADERS = ["엔진 모델명", "RPM", "출력(kW)", "토크(Nm)"]
ENGINE_KEYS = [s[0] for s in SCHEMA]

# ---------------------------------------------------------------- 샘플 엔진
# 제조사 3곳(가상명), 엔진 6개. 곡선 파라미터:
#   rpms: 곡선 RPM 포인트(8~12개), t_peak: 최대 토크(Nm), rpm_peak: 토크 피크 RPM,
#   floor: 저속측 토크 하한 비율 (포물선 폭은 정격 RPM까지 출력이 증가하도록 자동 산정)
ENGINES = [
    dict(
        manufacturer="대한파워텍", engine_model="DP34T",
        forklift_models="HF250D-9, HF300D-9, HF330D-9",
        rated_rpm=2200, low_idle=800, high_idle=2400,
        weight=305, etype="전자식", cyl=4, disp=3409, oil=8.5, cool=5.4,
        fuel=228, alt="24V 45A", starter=3.2,
        cert="Tier 4 Final / Stage V", after="DOC + DPF",
        rpms=[900, 1100, 1300, 1500, 1700, 1900, 2000, 2100, 2200],
        t_peak=270.0, rpm_peak=1500, floor=0.78,
    ),
    dict(
        manufacturer="대한파워텍", engine_model="DP24E",
        forklift_models="HF180D-9, HF200D-9",
        rated_rpm=2400, low_idle=780, high_idle=2600,
        weight=248, etype="전자식", cyl=4, disp=2392, oil=6.7, cool=4.5,
        fuel=235, alt="12V 60A", starter=2.7,
        cert="Tier 4 Final", after="DOC",
        rpms=[900, 1100, 1300, 1500, 1700, 1900, 2100, 2250, 2400],
        t_peak=185.0, rpm_peak=1600, floor=0.8,
    ),
    dict(
        manufacturer="CK엔진코리아", engine_model="QF3.8L",
        forklift_models="HF250D-9, HF350D-9, HF400D-9",
        rated_rpm=2300, low_idle=825, high_idle=2500,
        weight=331, etype="전자식", cyl=4, disp=3760, oil=9.5, cool=6.0,
        fuel=222, alt="24V 70A", starter=3.6,
        cert="Stage V", after="DOC + DPF + SCR",
        rpms=[800, 1000, 1200, 1400, 1500, 1600, 1800, 2000, 2100, 2200, 2300],
        t_peak=300.0, rpm_peak=1500, floor=0.75,
    ),
    dict(
        manufacturer="CK엔진코리아", engine_model="QF2.8M",
        forklift_models="HF150D-7, HF180D-7",
        rated_rpm=2500, low_idle=750, high_idle=2700,
        weight=214, etype="기계식", cyl=4, disp=2776, oil=6.0, cool=4.2,
        fuel=242, alt="12V 45A", starter=2.5,
        cert="Tier 3", after="없음",
        rpms=[900, 1100, 1300, 1500, 1700, 1900, 2100, 2300, 2400, 2500],
        t_peak=175.0, rpm_peak=1700, floor=0.82,
    ),
    dict(
        manufacturer="YS디젤", engine_model="4YS98T",
        forklift_models="HF200D-9, HF250D-9",
        rated_rpm=2400, low_idle=800, high_idle=2600,
        weight=238, etype="전자식", cyl=4, disp=3319, oil=7.4, cool=5.0,
        fuel=230, alt="12V 55A", starter=3.0,
        cert="Tier 4 Final / Stage V", after="DPF",
        rpms=[850, 1050, 1250, 1450, 1650, 1850, 2000, 2150, 2300, 2400],
        t_peak=250.0, rpm_peak=1560, floor=0.76,
    ),
    dict(
        manufacturer="YS디젤", engine_model="4YS88",
        forklift_models="HF150D-7, HF160D-7, HF180D-7",
        rated_rpm=2600, low_idle=730, high_idle=2800,
        weight=183, etype="기계식", cyl=4, disp=2189, oil=5.5, cool=3.8,
        fuel=248, alt="12V 40A", starter=2.3,
        cert="Tier 3", after="없음",
        rpms=[900, 1150, 1400, 1650, 1900, 2100, 2300, 2450, 2600],
        t_peak=140.0, rpm_peak=1650, floor=0.84,
    ),
]


def torque_at(rpm, t_peak, rpm_peak, width, floor):
    """토크 곡선: 피크 중심 포물선, 저속측은 floor 비율 아래로 내려가지 않음."""
    ratio = 1.0 - ((rpm - rpm_peak) / width) ** 2
    ratio = max(ratio, floor if rpm < rpm_peak else 0.55)
    return t_peak * ratio


def build_curve(e):
    """RPM 포인트별 (rpm, 출력 kW, 토크 Nm). 출력 = 토크 x RPM / 9549.

    포물선 폭은 3.0 x (정격 RPM - 토크 피크 RPM) 으로 잡는다 —
    이 폭이면 토크 감소가 완만해 정격 RPM까지 출력이 단조 증가한다
    (조건: 폭 > ~2.8 x 간격, 미분으로 유도).
    """
    width = 3.0 * (e["rated_rpm"] - e["rpm_peak"])
    pts = []
    for rpm in e["rpms"]:
        t = round(torque_at(rpm, e["t_peak"], e["rpm_peak"], width, e["floor"]), 1)
        p = round(t * rpm / 9549.0, 1)
        pts.append((rpm, p, t))
    return pts


def build_rows():
    """engines 행 + curves 행 생성. 정격/최대 값은 곡선에서 역산해 완전 일치시킨다."""
    engine_rows, curve_rows = [], []
    for e in ENGINES:
        pts = build_curve(e)
        by_rpm = {rpm: (p, t) for rpm, p, t in pts}
        rated_kw = by_rpm[e["rated_rpm"]][0]
        # 출력 최대점이 정격 RPM인지 검증 (샘플 데이터 무결성)
        assert rated_kw == max(p for _, p, _ in pts), e["engine_model"]
        max_t = max(t for _, _, t in pts)
        max_t_rpm = min(rpm for rpm, _, t in pts if t == max_t)
        engine_rows.append({
            "manufacturer": e["manufacturer"],
            "engine_model": e["engine_model"],
            "forklift_models": e["forklift_models"],
            "rated_power_kw": rated_kw,
            "rated_power_rpm": e["rated_rpm"],
            "max_torque_nm": max_t,
            "max_torque_rpm": max_t_rpm,
            "low_idle_rpm": e["low_idle"],
            "high_idle_rpm": e["high_idle"],
            "weight_kg": e["weight"],
            "engine_type": e["etype"],
            "cylinders": e["cyl"],
            "displacement_cc": e["disp"],
            "oil_capacity_l": e["oil"],
            "coolant_capacity_l": e["cool"],
            "fuel_consumption": e["fuel"],
            "alternator": e["alt"],
            "starter_motor_kw": e["starter"],
            "emission_cert": e["cert"],
            "aftertreatment": e["after"],
        })
        for rpm, p, t in pts:
            curve_rows.append((e["engine_model"], rpm, p, t))
    return engine_rows, curve_rows


def write_workbook(path, engine_rows, curve_rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "engines"
    ws.append(ENGINE_KEYS)
    for row in engine_rows:
        ws.append([row.get(k) for k in ENGINE_KEYS])

    ws2 = wb.create_sheet("curves")
    ws2.append(CURVE_HEADERS)
    for r in curve_rows:
        ws2.append(list(r))

    ws3 = wb.create_sheet("schema")
    ws3.append(SCHEMA_HEADERS)
    for s in SCHEMA:
        ws3.append(list(s))

    wb.save(path)
    print("생성:", os.path.relpath(path, BASE))


def write_sample_js(path, engine_rows, curve_rows):
    """js/sample-data.js — 웹페이지 '샘플 데이터 불러오기' 버튼용 내장 데이터."""
    curves = {}
    for model, rpm, p, t in curve_rows:
        curves.setdefault(model, []).append({"rpm": rpm, "power": p, "torque": t})
    data = {
        "schema": [
            {"key": s[0], "label": s[1], "group": s[2], "unit": s[3], "order": s[4]}
            for s in SCHEMA
        ],
        "engines": engine_rows,
        "curves": curves,
    }
    body = json.dumps(data, ensure_ascii=False, indent=2)
    js = (
        "/**\n"
        " * sample-data.js — 내장 샘플 데이터 (make_samples.py 가 자동 생성; 직접 수정 금지)\n"
        " * templates/엔진사양_샘플.xlsx 와 동일한 내용. '샘플 데이터 불러오기' 버튼에서 사용.\n"
        " */\n"
        "(function (root, factory) {\n"
        "  if (typeof module === \"object\" && module.exports) { module.exports = factory(); }\n"
        "  else { root.SAMPLE_DATA = factory(); }\n"
        "})(typeof self !== \"undefined\" ? self : this, function () {\n"
        "  \"use strict\";\n"
        "  return " + body.replace("\n", "\n  ") + ";\n"
        "});\n"
    )
    with open(path, "w", encoding="utf-8") as f:
        f.write(js)
    print("생성:", os.path.relpath(path, BASE))


def main():
    tdir = os.path.join(BASE, "templates")
    os.makedirs(tdir, exist_ok=True)
    engine_rows, curve_rows = build_rows()
    # 1) 빈 표준 양식 (헤더 + schema 만)
    write_workbook(os.path.join(tdir, "엔진사양_표준양식.xlsx"), [], [])
    # 2) 샘플 데이터
    write_workbook(os.path.join(tdir, "엔진사양_샘플.xlsx"), engine_rows, curve_rows)
    # 3) 내장 샘플 JS
    write_sample_js(os.path.join(BASE, "js", "sample-data.js"), engine_rows, curve_rows)


if __name__ == "__main__":
    main()
